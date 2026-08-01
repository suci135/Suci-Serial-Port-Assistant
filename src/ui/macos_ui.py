"""
macOS 原生风格串口调试工具 UI - 支持串口和蓝牙
严格遵循 macOS Human Interface Guidelines
"""

import json
import os
import sys
import time
from typing import List
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *
from datetime import datetime

from ..core.app_config import AppConfig
from ..core.quick_command_store import QuickCommandStore
from ..core.session_recorder import SessionRecorder
from ..core.command_sequence import CommandSequenceRunner
from ..core.payload_codec import analyze_payload
from ..core.receive_pause_buffer import ReceivePauseBuffer
from ..core.resources import resource_path as get_resource_path
from ..core.send_history import SendHistoryStore
from ..core.serial_manager import SerialManager, SerialDevice
from ..core.bluetooth_manager import BluetoothManager, BluetoothDevice
from .layout_metrics import (
    COMPOSER_MIN_HEIGHT,
    CONTENT_MARGIN,
    LEFT_SIDEBAR_WIDTH,
    RIGHT_SIDEBAR_WIDTH,
    SECTION_SPACING,
    SIDEBAR_HORIZONTAL_MARGIN,
    SIDEBAR_VERTICAL_MARGIN,
    SPLITTER_HANDLE_WIDTH,
    TOOLBAR_SPACING,
)
from .responsive_layout import layout_state_for_width
from .design import palette_for
from .workbench import AsciiTableWidget, RadixConverterWidget, SendComposer


def resource_path(relative_path: str) -> str:
    """获取资源文件的绝对路径，兼容 PyInstaller 打包环境"""
    return get_resource_path(relative_path)


class MessageContainer(QWidget):
    """带 hover 事件的消息容器"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._copy_btn = None
        self._copied_label = None

    def set_copy_widgets(self, copy_btn, copied_label):
        self._copy_btn = copy_btn
        self._copied_label = copied_label

    def enterEvent(self, event):
        if self._copy_btn and self._copied_label and not self._copied_label.isVisible():
            self._copy_btn.setVisible(True)
        super().enterEvent(event)

    def leaveEvent(self, event):
        if self._copy_btn:
            self._copy_btn.setVisible(False)
        super().leaveEvent(event)


class MacOSSerialUI(QWidget):
    """macOS 风格串口调试助手主界面"""

    data_sent = pyqtSignal(bytes)
    display_sent_signal = pyqtSignal(str, str)  # 新增信号：用于显示发送的数据
    manual_send_completed = pyqtSignal(bool, int)
    
    def __init__(self, config: AppConfig, serial_mgr: SerialManager):
        super().__init__()
        self.config = config
        self.serial_manager = serial_mgr
        self.serial_manager.set_auto_reconnect(
            self.config.get("serial.auto_reconnect", True)
        )
        self.bluetooth_manager = BluetoothManager()
        
        # 当前模式：'serial' 或 'bluetooth'
        self.current_mode = 'serial'
        
        # 统计数据
        self.sent_count = 0
        self.sent_bytes = 0
        self.received_count = 0
        self.received_bytes = 0
        self._rate_timestamp = time.monotonic()
        self._rate_sent_bytes = 0
        self._rate_received_bytes = 0
        self._send_rate = 0.0
        self._receive_rate = 0.0

        # 消息记录（用于导出）
        self.message_log = []  # [{"time": str, "direction": str, "content": str}]
        self.session_recorder = SessionRecorder()
        self.send_history = SendHistoryStore(
            self.config.config_dir / "send_history.json",
            self.config.get("send.history_limit", 100),
        )
        self.receive_pause_buffer = ReceivePauseBuffer(
            self.config.get("display.pause_buffer_bytes", 512 * 1024)
        )
        self._history_cursor = -1
        self._history_draft = ""
        self._restoring_history = False

        # 黑夜模式状态
        self.is_dark_mode = False
        
        # 配置文件路径
        self.config_file = "quick_commands.json"
        self.quick_command_store = QuickCommandStore(self.config_file)
        
        # 快捷输入数据 - 从配置文件加载
        self.quick_commands = self.load_quick_commands()
        
        # 连接动画定时器
        self.connecting_timer = QTimer()
        self.connecting_timer.timeout.connect(self.update_connecting_animation)
        self.connecting_dots = 0
        self.repeat_timer = QTimer(self)
        self.repeat_timer.timeout.connect(self.send_data)
        self.sequence_runner = CommandSequenceRunner(self)
        self.sequence_runner.command_sent.connect(self._on_sequence_command_sent)
        self.sequence_runner.command_result.connect(self._on_sequence_command_result)
        self.sequence_runner.completed.connect(self._on_sequence_completed)
        self.manual_send_completed.connect(self._on_manual_send_completed)
        
        self.init_ui()
        self.line_status_timer = QTimer(self)
        self.line_status_timer.setInterval(250)
        self.line_status_timer.timeout.connect(self.update_line_states)
        self.line_status_timer.start()
        self.apply_macos_style()
        self.connect_signals()
        self.refresh_devices()
    
    def init_ui(self):
        """Build the adaptive workbench shell."""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.workbench_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.workbench_splitter.setObjectName("workbenchSplitter")
        self.workbench_splitter.setHandleWidth(SPLITTER_HANDLE_WIDTH)
        self.workbench_splitter.setChildrenCollapsible(False)

        self.left_sidebar = self.create_left_sidebar()
        self.center_content = self.create_center_content()
        self.right_sidebar = self.create_right_sidebar()
        self.workbench_splitter.addWidget(self.left_sidebar)
        self.workbench_splitter.addWidget(self.center_content)
        self.workbench_splitter.addWidget(self.right_sidebar)
        self.workbench_splitter.setStretchFactor(0, 0)
        self.workbench_splitter.setStretchFactor(1, 1)
        self.workbench_splitter.setStretchFactor(2, 0)
        saved_sizes = self.config.get("window.splitter_sizes", [])
        initial_sizes = (
            saved_sizes
            if isinstance(saved_sizes, list) and len(saved_sizes) == 3
            else [LEFT_SIDEBAR_WIDTH, 720, RIGHT_SIDEBAR_WIDTH]
        )
        self.workbench_splitter.setSizes(initial_sizes)
        main_layout.addWidget(self.workbench_splitter)
        self.workbench_splitter.installEventFilter(self)

        # Persistent edge handles replace the toolbar's text panel buttons.
        # They remain available when either panel is collapsed.
        self.left_edge_button = QPushButton(self)
        self.left_edge_button.setObjectName("edgeToggleButton")
        self.left_edge_button.setFixedSize(24, 42)
        self.left_edge_button.setToolTip("收起连接侧栏")
        self.left_edge_button.clicked.connect(self.toggle_navigation_panel)

        self.right_edge_button = QPushButton(self)
        self.right_edge_button.setObjectName("edgeToggleButton")
        self.right_edge_button.setFixedSize(24, 42)
        self.right_edge_button.setToolTip("收起参数与命令面板")
        self.right_edge_button.clicked.connect(self.toggle_inspector_panel)
        self.left_edge_button.raise_()
        self.right_edge_button.raise_()

        self._navigation_requested = True
        self._inspector_requested = True
        self._update_edge_buttons()
        self._responsive_timer = QTimer(self)
        self._responsive_timer.setSingleShot(True)
        self._responsive_timer.setInterval(16)
        self._responsive_timer.timeout.connect(self._apply_responsive_layout)
        self._splitter_save_timer = QTimer(self)
        self._splitter_save_timer.setSingleShot(True)
        self._splitter_save_timer.setInterval(250)
        self._splitter_save_timer.timeout.connect(self._save_splitter_sizes)
        self.workbench_splitter.splitterMoved.connect(self._on_splitter_moved)
        QTimer.singleShot(0, self._apply_responsive_layout)

    def _save_splitter_sizes(self):
        sizes = self.workbench_splitter.sizes()
        if all(size > 0 for size in sizes):
            self.config.set("window.splitter_sizes", sizes)
            self.config.save_config()

    def _on_splitter_moved(self, _position, _index):
        """Keep edge handles attached to pane boundaries during dragging."""
        self._position_edge_buttons()
        self._splitter_save_timer.start()

    def resizeEvent(self, event):
        """Coalesce resize updates so dragging the window remains fluid."""
        if hasattr(self, "_responsive_timer"):
            self._responsive_timer.start()
        if hasattr(self, "left_edge_button"):
            QTimer.singleShot(0, self._position_edge_buttons)
        super().resizeEvent(event)

    def showEvent(self, event):
        """Re-anchor edge controls after the first native layout pass."""
        super().showEvent(event)
        QTimer.singleShot(0, self._position_edge_buttons)
        QTimer.singleShot(50, self._position_edge_buttons)

    def eventFilter(self, watched, event):
        if (
            hasattr(self, "workbench_splitter")
            and watched is self.workbench_splitter
            and event.type()
            in (QEvent.Type.Resize, QEvent.Type.Show, QEvent.Type.LayoutRequest)
        ):
            QTimer.singleShot(0, self._position_edge_buttons)
        return super().eventFilter(watched, event)

    def _apply_responsive_layout(self):
        if not hasattr(self, "workbench_splitter"):
            return
        state = layout_state_for_width(self.width())
        show_navigation = state.show_navigation and self._navigation_requested
        show_inspector = state.show_inspector and self._inspector_requested
        self.left_sidebar.setVisible(show_navigation)
        self.right_sidebar.setVisible(show_inspector)
        self._update_edge_buttons()
        self.search_input.setMaximumWidth(112 if state.compact_toolbar else 180)
        self.timestamp_check.setVisible(not state.compact_toolbar)
        self.autoscroll_check.setVisible(not state.compact_toolbar)
        self.export_btn.setVisible(not state.compact_toolbar)

    def _position_edge_buttons(self):
        """Place collapse handles on the current pane boundaries."""
        if not hasattr(self, "left_edge_button"):
            return
        y = max(8, (self.height() - self.left_edge_button.height()) // 2)
        left_x = (
            self.left_sidebar.mapTo(self, QPoint(self.left_sidebar.width(), 0)).x()
            - self.left_edge_button.width() // 2
            if self.left_sidebar.isVisible()
            else 0
        )
        right_x = (
            self.right_sidebar.mapTo(self, QPoint(0, 0)).x()
            - self.right_edge_button.width() // 2
            if self.right_sidebar.isVisible()
            else self.width() - self.right_edge_button.width()
        )
        self.left_edge_button.move(max(0, left_x), y)
        self.right_edge_button.move(
            min(self.width() - self.right_edge_button.width(), right_x), y
        )
        self.left_edge_button.raise_()
        self.right_edge_button.raise_()

    def _update_edge_buttons(self):
        left_open = self.left_sidebar.isVisible()
        right_open = self.right_sidebar.isVisible()
        self.left_edge_button.setIcon(
            self.style().standardIcon(
                QStyle.StandardPixmap.SP_ArrowLeft
                if left_open else QStyle.StandardPixmap.SP_ArrowRight
            )
        )
        self.left_edge_button.setIconSize(QSize(12, 12))
        self.left_edge_button.setToolTip("收起连接侧栏" if left_open else "展开连接侧栏")
        self.right_edge_button.setIcon(
            self.style().standardIcon(
                QStyle.StandardPixmap.SP_ArrowRight
                if right_open else QStyle.StandardPixmap.SP_ArrowLeft
            )
        )
        self.right_edge_button.setIconSize(QSize(12, 12))
        self.right_edge_button.setToolTip(
            "收起参数与命令面板" if right_open else "展开参数与命令面板"
        )
        QTimer.singleShot(0, self._position_edge_buttons)

    def toggle_navigation_panel(self):
        show = not self.left_sidebar.isVisible()
        self._navigation_requested = show
        self.left_sidebar.setVisible(show)
        self._update_edge_buttons()

    def toggle_inspector_panel(self):
        show = not self.right_sidebar.isVisible()
        self._inspector_requested = show
        self.right_sidebar.setVisible(show)
        self._update_edge_buttons()

    def create_left_sidebar(self):
        """创建左侧边栏 - 串口/蓝牙连接设置"""
        sidebar = QFrame()
        sidebar.setObjectName("leftSidebar")
        sidebar.setMinimumWidth(180)
        sidebar.setMaximumWidth(280)
        
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(
            SIDEBAR_HORIZONTAL_MARGIN,
            SIDEBAR_VERTICAL_MARGIN,
            SIDEBAR_HORIZONTAL_MARGIN,
            SIDEBAR_VERTICAL_MARGIN,
        )
        layout.setSpacing(SECTION_SPACING)
        
        # 模式选择
        mode_label = QLabel("模式")
        mode_label.setObjectName("compactLabel")
        layout.addWidget(mode_label)
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["串口", "蓝牙"])
        self.mode_combo.currentTextChanged.connect(self.on_mode_changed)
        layout.addWidget(self.mode_combo)
        
        # 设备选择
        device_label = QLabel("设备")
        device_label.setObjectName("compactLabel")
        layout.addWidget(device_label)
        
        # 设备选择容器（包含下拉框和刷新按钮）
        device_container = QHBoxLayout()
        device_container.setSpacing(5)
        
        self.port_combo = QComboBox()
        device_container.addWidget(self.port_combo, 1)
        
        # 刷新/扫描按钮
        self.refresh_btn = QPushButton("⟳")
        self.refresh_btn.setObjectName("refreshButton")
        self.refresh_btn.setFixedSize(32, 32)
        self.refresh_btn.setToolTip("刷新设备列表")
        self.refresh_btn.clicked.connect(self.refresh_devices)
        self.refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        device_container.addWidget(self.refresh_btn)
        
        layout.addLayout(device_container)
        
        # 串口配置区域（容器）
        self.serial_config_widget = QWidget()
        serial_config_layout = QVBoxLayout(self.serial_config_widget)
        serial_config_layout.setContentsMargins(0, 0, 0, 0)
        serial_config_layout.setSpacing(7)
        
        # Baud Rate
        baud_label = QLabel("波特率")
        baud_label.setObjectName("compactLabel")
        serial_config_layout.addWidget(baud_label)
        self.baud_combo = QComboBox()
        self.baud_combo.setEditable(True)
        self.baud_combo.addItems(["9600", "19200", "38400", "57600", "115200"])
        self.baud_combo.setCurrentText(str(self.config.get("serial.baud_rate", 9600)))
        serial_config_layout.addWidget(self.baud_combo)
        
        # Data Bits
        data_label = QLabel("数据位")
        data_label.setObjectName("compactLabel")
        serial_config_layout.addWidget(data_label)
        self.data_bits_combo = QComboBox()
        self.data_bits_combo.addItems(["5", "6", "7", "8"])
        self.data_bits_combo.setCurrentText(str(self.config.get("serial.data_bits", 8)))
        serial_config_layout.addWidget(self.data_bits_combo)
        
        # Stop Bits
        stop_label = QLabel("停止位")
        stop_label.setObjectName("compactLabel")
        serial_config_layout.addWidget(stop_label)
        self.stop_bits_combo = QComboBox()
        self.stop_bits_combo.addItems(["1", "1.5", "2"])
        self.stop_bits_combo.setCurrentText(str(self.config.get("serial.stop_bits", 1)))
        serial_config_layout.addWidget(self.stop_bits_combo)
        
        # Parity
        parity_label = QLabel("校验位")
        parity_label.setObjectName("compactLabel")
        serial_config_layout.addWidget(parity_label)
        self.parity_combo = QComboBox()
        self.parity_combo.addItems(["None", "Even", "Odd", "Mark", "Space"])
        self.parity_combo.setCurrentText(self.config.get("serial.parity", "None"))
        serial_config_layout.addWidget(self.parity_combo)

        flow_label = QLabel("流控制")
        flow_label.setObjectName("compactLabel")
        serial_config_layout.addWidget(flow_label)
        self.flow_control_combo = QComboBox()
        self.flow_control_combo.addItems(["None", "XON/XOFF", "RTS/CTS", "DSR/DTR"])
        self.flow_control_combo.setCurrentText(
            self.config.get("serial.flow_control", "None")
        )
        serial_config_layout.addWidget(self.flow_control_combo)

        self.auto_reconnect_check = QCheckBox("意外断开后自动重连")
        self.auto_reconnect_check.setChecked(
            self.config.get("serial.auto_reconnect", True)
        )
        self.auto_reconnect_check.toggled.connect(self.set_auto_reconnect)
        serial_config_layout.addWidget(self.auto_reconnect_check)
        
        layout.addWidget(self.serial_config_widget)
        
        # 蓝牙配置区域（容器）
        self.bluetooth_config_widget = QWidget()
        bluetooth_config_layout = QVBoxLayout(self.bluetooth_config_widget)
        bluetooth_config_layout.setContentsMargins(0, 0, 0, 0)
        bluetooth_config_layout.setSpacing(7)
        
        # PyBluez 配置（经典蓝牙）
        self.pybluez_config_widget = QWidget()
        pybluez_config_layout = QVBoxLayout(self.pybluez_config_widget)
        pybluez_config_layout.setContentsMargins(0, 0, 0, 0)
        pybluez_config_layout.setSpacing(7)
        
        # RFCOMM端口
        port_label = QLabel("RFCOMM端口")
        port_label.setObjectName("compactLabel")
        pybluez_config_layout.addWidget(port_label)
        self.rfcomm_port_spin = QSpinBox()
        self.rfcomm_port_spin.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.rfcomm_port_spin.setRange(1, 30)
        self.rfcomm_port_spin.setValue(1)
        pybluez_config_layout.addWidget(self.rfcomm_port_spin)
        
        bluetooth_config_layout.addWidget(self.pybluez_config_widget)
        
        # 蓝牙提示信息
        bt_info = QLabel()
        bt_info.setObjectName("infoLabel")
        bt_info.setWordWrap(True)
        self.bt_info_label = bt_info
        bluetooth_config_layout.addWidget(bt_info)
        
        layout.addWidget(self.bluetooth_config_widget)
        
        # 默认隐藏蓝牙配置
        self.bluetooth_config_widget.hide()
        
        # 连接按钮
        self.connect_btn = QPushButton("连接")
        self.connect_btn.setObjectName("primaryButton")
        self.connect_btn.setMinimumHeight(32)
        self.connect_btn.clicked.connect(self.toggle_connection)
        layout.addWidget(self.connect_btn)
        
        # 状态指示
        status_frame = QFrame()
        status_frame.setObjectName("statusFrame")
        status_layout = QHBoxLayout(status_frame)
        status_layout.setContentsMargins(8, 6, 8, 6)
        
        self.status_indicator = QLabel("●")
        self.status_indicator.setObjectName("statusDot")
        status_layout.addWidget(self.status_indicator)
        
        self.status_label = QLabel("未连接")
        self.status_label.setObjectName("statusText")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        
        layout.addWidget(status_frame)

        # 串口线路监控与手动控制
        self.line_control_card = QFrame()
        self.line_control_card.setObjectName("lineControlCard")
        line_card_layout = QVBoxLayout(self.line_control_card)
        line_card_layout.setContentsMargins(10, 9, 10, 9)
        line_card_layout.setSpacing(7)

        line_title = QLabel("线路监控")
        line_title.setObjectName("sectionTitle")
        line_card_layout.addWidget(line_title)

        self.connection_summary = QLabel("9600 · 8N1 · 无流控")
        self.connection_summary.setObjectName("connectionSummary")
        line_card_layout.addWidget(self.connection_summary)

        state_row = QHBoxLayout()
        state_row.setSpacing(5)
        self.line_state_labels = {}
        for name in ("CTS", "DSR", "DCD", "RI"):
            label = QLabel(f"● {name}")
            label.setObjectName("lineStateInactive")
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.line_state_labels[name] = label
            state_row.addWidget(label, 1)
        line_card_layout.addLayout(state_row)

        control_row = QHBoxLayout()
        control_row.setSpacing(6)
        self.dtr_button = QPushButton("DTR")
        self.dtr_button.setObjectName("lineControlButton")
        self.dtr_button.setCheckable(True)
        self.dtr_button.setToolTip("切换串口 DTR 输出线")
        self.dtr_button.toggled.connect(self.set_dtr_line)
        self.rts_button = QPushButton("RTS")
        self.rts_button.setObjectName("lineControlButton")
        self.rts_button.setCheckable(True)
        self.rts_button.setToolTip("切换串口 RTS 输出线")
        self.rts_button.toggled.connect(self.set_rts_line)
        self.break_button = QPushButton("BREAK")
        self.break_button.setObjectName("lineControlButton")
        self.break_button.setToolTip("发送 250 ms BREAK 信号")
        self.break_button.clicked.connect(self.send_break_signal)
        for button in (self.dtr_button, self.rts_button, self.break_button):
            button.setEnabled(False)
            control_row.addWidget(button)
        line_card_layout.addLayout(control_row)
        layout.addWidget(self.line_control_card)
        layout.addStretch()
        
        return sidebar

    def create_center_content(self):
        """创建中央内容区"""
        content = QFrame()
        content.setObjectName("centerContent")
        
        layout = QVBoxLayout(content)
        layout.setContentsMargins(CONTENT_MARGIN, CONTENT_MARGIN, CONTENT_MARGIN, 10)
        layout.setSpacing(10)
        
        # 顶部工具栏
        toolbar = QHBoxLayout()
        toolbar.setSpacing(TOOLBAR_SPACING)

        # 格式选择
        toolbar.addWidget(QLabel("格式:"))
        self.format_combo = QComboBox()
        self.format_combo.addItems(["HEX", "ASCII", "UTF-8"])
        self.format_combo.setCurrentText("UTF-8")
        self.format_combo.setMinimumWidth(100)
        self.format_combo.setMaximumWidth(130)
        toolbar.addWidget(self.format_combo)
        
        toolbar.addSpacing(8)
        
        # 时间戳
        self.timestamp_check = QCheckBox("显示时间戳")
        self.timestamp_check.setChecked(True)
        self.timestamp_check.stateChanged.connect(self.toggle_timestamps)
        toolbar.addWidget(self.timestamp_check)
        
        # 自动滚动
        self.autoscroll_check = QCheckBox("自动滚动")
        self.autoscroll_check.setChecked(True)
        toolbar.addWidget(self.autoscroll_check)

        # 消息搜索
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索消息...")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.setMaximumWidth(140)
        self.search_input.textChanged.connect(self.filter_messages)
        toolbar.addWidget(self.search_input)
        
        toolbar.addStretch()

        self.pause_display_btn = QPushButton("暂停显示")
        self.pause_display_btn.setObjectName("toolbarButton")
        self.pause_display_btn.setCheckable(True)
        self.pause_display_btn.setToolTip("暂停界面刷新，但继续接收并记录数据")
        self.pause_display_btn.toggled.connect(self.toggle_receive_pause)
        toolbar.addWidget(self.pause_display_btn)

        # 清除按钮
        self.clear_btn = QPushButton("清除")
        self.clear_btn.setObjectName("secondaryButton")
        self.clear_btn.clicked.connect(self.clear_display)
        toolbar.addWidget(self.clear_btn)

        # 会话录制
        self.record_btn = QPushButton("录制")
        self.record_btn.setObjectName("secondaryButton")
        self.record_btn.setToolTip("将收发原始数据保存为 CSV")
        self.record_btn.clicked.connect(self.toggle_recording)
        toolbar.addWidget(self.record_btn)

        # 导出按钮
        self.export_btn = QPushButton("导出 Excel")
        self.export_btn.setObjectName("secondaryButton")
        self.export_btn.clicked.connect(self.export_to_excel)
        toolbar.addWidget(self.export_btn)

        
        layout.addLayout(toolbar)
        
        # 数据显示区 - 使用 QScrollArea + QWidget 实现气泡效果
        display_container = QFrame()
        display_container.setObjectName("dataDisplayContainer")
        display_container_layout = QVBoxLayout(display_container)
        display_container_layout.setContentsMargins(0, 0, 0, 0)
        display_container_layout.setSpacing(0)
        
        # 创建滚动区域
        self.scroll_area = QScrollArea()
        self.scroll_area.setObjectName("chatScrollArea")
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        
        # 创建消息容器
        self.messages_widget = QWidget()
        self.messages_widget.setObjectName("messagesWidget")
        self.messages_layout = QVBoxLayout(self.messages_widget)
        self.messages_layout.setContentsMargins(8, 8, 8, 8)
        self.messages_layout.setSpacing(6)
        self.messages_layout.addStretch()
        
        self.scroll_area.setWidget(self.messages_widget)
        display_container_layout.addWidget(self.scroll_area)
        
        layout.addWidget(display_container, 1)

        layout.addWidget(self.create_send_composer())
        
        # 底部状态栏
        status_bar = QHBoxLayout()
        status_bar.setSpacing(12)
        
        self.sent_stats = QLabel("发送: 0 (0 字节)")
        self.sent_stats.setObjectName("statsLabel")
        status_bar.addWidget(self.sent_stats)
        
        self.received_stats = QLabel("接收: 0 (0 字节)")
        self.received_stats.setObjectName("statsLabel")
        status_bar.addWidget(self.received_stats)

        self.rate_stats = QLabel("速率: ↓ 0 B/s  ↑ 0 B/s")
        self.rate_stats.setObjectName("statsLabel")
        status_bar.addWidget(self.rate_stats)
        
        status_bar.addStretch()
        
        layout.addLayout(status_bar)
        
        return content

    def create_send_composer(self):
        """Create the always-nearby send area below the terminal."""
        composer = SendComposer(COMPOSER_MIN_HEIGHT)
        composer.send_requested.connect(self.send_data)
        composer.repeat_toggled.connect(self.toggle_repeat_send)
        composer.interval_changed.connect(self.update_repeat_interval)
        composer.history_requested.connect(self.show_send_history)
        composer.history_previous.connect(lambda: self.recall_send_history(1))
        composer.history_next.connect(lambda: self.recall_send_history(-1))

        # Compatibility aliases keep the existing communication logic stable.
        self.send_format_combo = composer.format_combo
        self.send_input = composer.input
        self.send_btn = composer.send_button
        self.send_history_btn = composer.history_button
        self.payload_status = composer.payload_status
        self.add_carriage = composer.add_carriage
        self.add_newline = composer.add_newline
        self.repeat_check = composer.repeat_check
        self.repeat_interval_spin = composer.interval_spin
        self.send_input.textChanged.connect(self._on_send_input_changed)
        self.send_format_combo.currentTextChanged.connect(self.update_send_analysis)
        self.add_carriage.toggled.connect(self.update_send_analysis)
        self.add_newline.toggled.connect(self.update_send_analysis)
        QTimer.singleShot(0, self.update_send_analysis)
        return composer

    def _on_send_input_changed(self):
        if not self._restoring_history:
            self._history_cursor = -1
            self._history_draft = self.send_input.toPlainText()
        self.update_send_analysis()

    def update_send_analysis(self):
        analysis = analyze_payload(
            self.send_input.toPlainText(),
            self.send_format_combo.currentText(),
            self.add_carriage.isChecked(),
            self.add_newline.isChecked(),
        )
        if analysis.valid:
            self.payload_status.setObjectName("payloadStatus")
            self.payload_status.setText(f"{analysis.byte_count} 字节")
            self.payload_status.setToolTip("实际发送的字节数（包含行尾符）")
        else:
            self.payload_status.setObjectName("payloadStatusError")
            self.payload_status.setText("格式错误")
            self.payload_status.setToolTip(analysis.error)
        self.payload_status.style().unpolish(self.payload_status)
        self.payload_status.style().polish(self.payload_status)

    def _restore_history_entry(self, entry):
        self._restoring_history = True
        self.send_format_combo.setCurrentText(entry["format"])
        self.send_input.setPlainText(entry["text"])
        self.send_input.moveCursor(QTextCursor.MoveOperation.End)
        self._restoring_history = False
        self.update_send_analysis()
        self.send_input.setFocus()

    def recall_send_history(self, direction: int):
        entries = self.send_history.recent()
        if not entries:
            return
        if direction > 0:
            if self._history_cursor < 0:
                self._history_draft = self.send_input.toPlainText()
            self._history_cursor = min(self._history_cursor + 1, len(entries) - 1)
            self._restore_history_entry(entries[self._history_cursor])
        elif self._history_cursor > 0:
            self._history_cursor -= 1
            self._restore_history_entry(entries[self._history_cursor])
        elif self._history_cursor == 0:
            self._history_cursor = -1
            self._restoring_history = True
            self.send_input.setPlainText(self._history_draft)
            self._restoring_history = False
            self.update_send_analysis()

    def show_send_history(self):
        menu = QMenu(self)
        entries = self.send_history.recent()
        if not entries:
            empty = menu.addAction("暂无发送历史")
            empty.setEnabled(False)
        for entry in entries[:20]:
            preview = " ".join(entry["text"].split())
            if len(preview) > 42:
                preview = preview[:42] + "…"
            action = menu.addAction(f'[{entry["format"]}]  {preview}')
            action.triggered.connect(
                lambda _checked=False, item=entry: self._restore_history_entry(item)
            )
        if entries:
            menu.addSeparator()
            clear_action = menu.addAction("清空发送历史")
            clear_action.triggered.connect(self.send_history.clear)
        position = self.send_history_btn.mapToGlobal(
            QPoint(0, self.send_history_btn.height())
        )
        menu.exec(position)

    def create_right_sidebar(self):
        """Create the context inspector for parameters and commands."""
        sidebar = QFrame()
        sidebar.setObjectName("rightSidebar")
        sidebar.setMinimumWidth(280)
        sidebar.setMaximumWidth(420)
        
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(12, SIDEBAR_VERTICAL_MARGIN, 12, 10)
        layout.setSpacing(SECTION_SPACING)
        
        tabs = QTabWidget()
        tabs.setObjectName("inspectorTabs")
        tabs.setDocumentMode(False)
        tabs.tabBar().setObjectName("inspectorTabBar")
        tabs.tabBar().setDrawBase(False)
        tabs.tabBar().setExpanding(False)
        self.inspector_tabs = tabs

        parameters_tab = QWidget()
        parameters_tab.setObjectName("inspectorPage")
        parameters_layout = QVBoxLayout(parameters_tab)
        parameters_layout.setContentsMargins(4, 10, 4, 4)
        parameters_layout.setSpacing(SECTION_SPACING)
        parameters_title = QLabel("连接参数")
        parameters_title.setObjectName("sectionTitle")
        parameters_layout.addWidget(parameters_title)
        parameters_layout.addWidget(self.serial_config_widget)
        parameters_layout.addWidget(self.bluetooth_config_widget)
        parameters_layout.addStretch()
        tabs.addTab(parameters_tab, "参数")

        commands_tab = QWidget()
        commands_tab.setObjectName("inspectorPage")
        commands_layout = QVBoxLayout(commands_tab)
        commands_layout.setContentsMargins(4, 10, 4, 4)
        commands_layout.setSpacing(SECTION_SPACING)
        quick_title = QLabel("快捷命令")
        quick_title.setObjectName("sectionTitle")
        commands_layout.addWidget(quick_title)
        self.create_quick_input_panel(commands_layout)
        tabs.addTab(commands_tab, "命令")

        self.ascii_table = AsciiTableWidget()
        self.ascii_table.insert_text_requested.connect(self.insert_ascii_text)
        self.ascii_table.insert_hex_requested.connect(self.insert_ascii_hex)
        tabs.addTab(self.ascii_table, "ASCII")

        self.radix_converter = RadixConverterWidget()
        self.radix_converter.insert_hex_requested.connect(self.insert_ascii_hex)
        tabs.addTab(self.radix_converter, "进制")

        layout.addWidget(tabs, 1)
        
        # 底部按钮区域
        bottom_buttons = QHBoxLayout()
        bottom_buttons.setSpacing(8)
        bottom_buttons.addStretch()
        
        # GitHub 按钮
        self.github_btn = QPushButton()
        self.github_btn.setObjectName("githubButton")
        self.github_btn.setFixedSize(32, 32)
        self.github_btn.setToolTip("访问 GitHub 仓库")
        self.github_btn.clicked.connect(self.open_github)
        self.github_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        
        # 设置 GitHub 图标
        from PyQt6.QtGui import QIcon
        self.github_btn.setIcon(QIcon(resource_path("src/resource/GitHub.png")))
        self.github_btn.setIconSize(QSize(20, 20))
        
        # 黑夜模式按钮
        self.dark_mode_btn = QPushButton()
        self.dark_mode_btn.setObjectName("darkModeButton")
        self.dark_mode_btn.setFixedSize(32, 32)
        self.dark_mode_btn.setToolTip("切换黑夜模式")
        self.dark_mode_btn.clicked.connect(self.toggle_dark_mode)
        self.dark_mode_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        
        # 设置黑夜模式图标
        self.dark_mode_btn.setIcon(QIcon(resource_path("src/resource/dark.png")))
        self.dark_mode_btn.setIconSize(QSize(20, 20))
        
        bottom_buttons.addWidget(self.github_btn)
        bottom_buttons.addWidget(self.dark_mode_btn)
        layout.addLayout(bottom_buttons)
        
        return sidebar

    def insert_ascii_text(self, text: str):
        """Insert a printable ASCII character using the active send format."""
        if self.send_format_combo.currentText() == "HEX":
            self._append_hex_bytes(text.encode("ascii"))
        else:
            self.send_input.insertPlainText(text)
            self.send_input.setFocus()

    def insert_ascii_hex(self, hexadecimal: str):
        """Convert existing input losslessly before switching to HEX."""
        existing = self.send_input.toPlainText()
        current_format = self.send_format_combo.currentText()
        prefix = b""
        if existing and current_format != "HEX":
            analysis = analyze_payload(existing, current_format)
            if not analysis.valid:
                self.on_error(analysis.error)
                return
            prefix = analysis.payload
        elif existing:
            analysis = analyze_payload(existing, "HEX")
            if not analysis.valid:
                self.on_error(analysis.error)
                return
            prefix = analysis.payload
        self.send_format_combo.setCurrentText("HEX")
        self.send_input.setPlainText(
            " ".join(f"{byte:02X}" for byte in prefix + bytes.fromhex(hexadecimal))
        )
        self.send_input.moveCursor(QTextCursor.MoveOperation.End)
        self.send_input.setFocus()

    def _append_hex_bytes(self, payload: bytes):
        existing = analyze_payload(self.send_input.toPlainText(), "HEX")
        if not existing.valid:
            self.on_error(existing.error)
            return
        combined = existing.payload + payload
        self.send_input.setPlainText(" ".join(f"{byte:02X}" for byte in combined))
        self.send_input.moveCursor(QTextCursor.MoveOperation.End)
        self.send_input.setFocus()
    
    def create_quick_input_panel(self, parent_layout):
        """创建快捷输入面板"""
        # 面板标题和表头
        header_frame = QFrame()
        header_frame.setObjectName("quickInputHeader")
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(6, 4, 6, 4)
        header_layout.setSpacing(6)
        
        # 表头标签
        content_label = QLabel("内容")
        content_label.setObjectName("headerLabel")
        header_layout.addWidget(content_label, 1)
        
        action_label = QLabel("操作")
        action_label.setObjectName("headerLabel")
        action_label.setFixedWidth(94)
        action_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(action_label)
        
        parent_layout.addWidget(header_frame)
        
        # 滚动区域
        scroll_area = QScrollArea()
        scroll_area.setObjectName("quickInputScrollArea")
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setMinimumHeight(160)
        scroll_area.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # 命令列表容器
        commands_widget = QWidget()
        commands_widget.setObjectName("quickInputCommands")
        self.commands_layout = QVBoxLayout(commands_widget)
        self.commands_layout.setContentsMargins(0, 0, 0, 0)
        self.commands_layout.setSpacing(2)
        
        # 创建命令行
        self.command_rows = []
        self.commands_layout.addStretch()
        for i, cmd in enumerate(self.quick_commands):
            self.create_command_row(cmd, i)
        scroll_area.setWidget(commands_widget)
        parent_layout.addWidget(scroll_area)
        
        # 添加新命令按钮
        add_btn = QPushButton("+ 添加命令")
        add_btn.setObjectName("addCommandButton")
        add_btn.clicked.connect(self.add_new_command)
        parent_layout.addWidget(add_btn)

        sequence_layout = QHBoxLayout()
        self.run_sequence_btn = QPushButton("执行启用命令")
        self.run_sequence_btn.setObjectName("secondaryButton")
        self.run_sequence_btn.setToolTip("按命令列表中的启用状态和延时依次发送")
        self.run_sequence_btn.clicked.connect(self.run_enabled_commands)
        self.stop_sequence_btn = QPushButton("停止")
        self.stop_sequence_btn.setObjectName("secondaryButton")
        self.stop_sequence_btn.setEnabled(False)
        self.stop_sequence_btn.clicked.connect(self.stop_command_sequence)
        sequence_layout.addWidget(self.run_sequence_btn)
        sequence_layout.addWidget(self.stop_sequence_btn)
        parent_layout.addLayout(sequence_layout)
    
    def create_command_row(self, cmd_data, index):
        """创建命令行"""
        row_frame = QFrame()
        row_frame.setObjectName("commandRow")
        row_layout = QHBoxLayout(row_frame)
        row_layout.setContentsMargins(3, 3, 3, 3)
        row_layout.setSpacing(3)
        
        # 命令内容（可编辑）- 现在占据更多空间
        content_container = QFrame()
        content_container.setObjectName("commandContentContainer")
        content_container_layout = QVBoxLayout(content_container)
        content_container_layout.setContentsMargins(8, 3, 8, 3)
        content_container_layout.setSpacing(2)
        
        command_edit = QLineEdit(cmd_data["command"])
        command_edit.setObjectName("commandEdit")
        command_edit.setPlaceholderText("输入命令...")
        command_edit.textChanged.connect(lambda text, idx=index: self.on_command_changed(idx, text))
        content_container_layout.addWidget(command_edit)
        
        # 描述现在也可以编辑
        desc_edit = QLineEdit(cmd_data["description"])
        desc_edit.setObjectName("commandDescriptionEdit")
        desc_edit.setPlaceholderText("输入描述...")
        desc_edit.textChanged.connect(lambda text, idx=index: self.on_description_changed(idx, text))
        content_container_layout.addWidget(desc_edit)
        
        row_layout.addWidget(content_container, 1)
        
        # 操作按钮
        action_layout = QHBoxLayout()
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.setSpacing(4)
        
        send_btn = QPushButton("发送")
        send_btn.setObjectName("quickSendButton")
        send_btn.setFixedSize(42, 28)
        send_btn.clicked.connect(lambda checked, idx=index: self.send_quick_command(idx))
        action_layout.addWidget(send_btn)
        
        delete_btn = QPushButton("删除")
        delete_btn.setObjectName("deleteButton")
        delete_btn.setFixedSize(42, 28)
        delete_btn.clicked.connect(lambda checked, idx=index: self.delete_command(idx))
        action_layout.addWidget(delete_btn)
        
        action_widget = QWidget()
        action_widget.setLayout(action_layout)
        action_widget.setFixedWidth(94)
        row_layout.addWidget(action_widget)
        
        # 保存行数据
        row_data = {
            'frame': row_frame,
            'command_edit': command_edit,
            'desc_edit': desc_edit,  # 改为可编辑的输入框
            'send_btn': send_btn,
            'delete_btn': delete_btn
        }
        self.command_rows.append(row_data)
        # Keep the expanding spacer at the end. Rebuilding after a deletion must
        # never place rows below it, otherwise a blank block appears at the top.
        self.commands_layout.insertWidget(
            max(0, self.commands_layout.count() - 1), row_frame
        )
    
    def on_command_changed(self, index, text):
        """命令内容改变"""
        if index < len(self.quick_commands):
            self.quick_commands[index]["command"] = text
            self.save_quick_commands()  # 自动保存
    
    def on_description_changed(self, index, text):
        """描述内容改变"""
        if index < len(self.quick_commands):
            self.quick_commands[index]["description"] = text
            self.save_quick_commands()  # 自动保存
    
    def on_hex_changed(self, index, state):
        """HEX状态改变"""
        if index < len(self.quick_commands):
            self.quick_commands[index]["is_hex"] = state == Qt.CheckState.Checked.value
    
    def on_delay_changed(self, index, text):
        """延时改变"""
        if index < len(self.quick_commands):
            try:
                delay = int(text) if text else 1000
                self.quick_commands[index]["delay"] = delay
            except ValueError:
                pass
    
    def send_quick_command(self, index):
        """发送快捷命令"""
        # 获取当前管理器
        manager = self.serial_manager if self.current_mode == 'serial' else self.bluetooth_manager
        
        if not manager.is_connected:
            self.on_error("设备未连接")
            return
        
        if index >= len(self.quick_commands):
            return
        
        cmd_data = self.quick_commands[index]
        command = cmd_data["command"].strip()
        if not command:
            return
        
        # 使用发送控制面板选择的格式
        format_type = self.send_format_combo.currentText()
        
        # 先在UI显示发送的数据
        self.display_sent_data(command, format_type)
        
        # 在后台线程中发送，避免阻塞 UI
        import threading
        def send_thread():
            success = False
            
            if format_type == "HEX":
                success = manager.send_hex_string(command)
            elif format_type == "ASCII":
                success = manager.send_text(command, 'ascii')
            else:  # UTF-8
                success = manager.send_text(command, 'utf-8')
            
            if success:
                # 更新统计
                self.sent_count += 1
                self.sent_bytes += len(command.encode('utf-8'))
        
        thread = threading.Thread(target=send_thread, daemon=True)
        thread.start()
        
        # 立即更新统计
        self.update_stats()
    
    def add_new_command(self):
        """添加新命令"""
        new_cmd = {
            "name": f"自定义命令{len(self.quick_commands) + 1}",
            "command": "",
            "description": "自定义命令",
            "enabled": True,
            "delay": 1000,
            "is_hex": False
        }
        self.quick_commands.append(new_cmd)
        self.create_command_row(new_cmd, len(self.quick_commands) - 1)
        self.save_quick_commands()  # 保存配置
    
    def delete_command(self, index):
        """删除命令"""
        if index < len(self.quick_commands) and index < len(self.command_rows):
            # 移除UI
            row_data = self.command_rows[index]
            self.commands_layout.removeWidget(row_data['frame'])
            row_data['frame'].deleteLater()
            
            # 移除数据
            self.quick_commands.pop(index)
            self.command_rows.pop(index)
            
            # 重新索引剩余的命令
            self.refresh_command_indices()
            self.save_quick_commands()  # 保存配置
    
    def refresh_command_indices(self):
        """刷新命令索引"""
        # 重新创建所有命令行以更新索引
        for row_data in self.command_rows:
            self.commands_layout.removeWidget(row_data['frame'])
            row_data['frame'].deleteLater()
        
        self.command_rows.clear()
        for i, cmd in enumerate(self.quick_commands):
            self.create_command_row(cmd, i)

    def apply_macos_style(self):
        """应用 macOS 风格样式"""
        # 加载字体
        from PyQt6.QtGui import QFontDatabase
        font_id = QFontDatabase.addApplicationFont(resource_path("src/resource/AlimamaFangYuanTiVF-Thin-2.ttf"))
        if font_id != -1:
            font_families = QFontDatabase.applicationFontFamilies(font_id)
            if font_families:
                font_family = font_families[0]
            else:
                font_family = "Microsoft YaHei"
        else:
            font_family = "Microsoft YaHei"
        
        palette = palette_for(self.is_dark_mode)
        bg_color = palette.surface
        sidebar_bg = palette.sidebar
        content_bg = palette.content
        input_bg = palette.input
        text_color = palette.text
        secondary_text = palette.secondary_text
        tertiary_text = palette.tertiary_text
        border_color = palette.border
        display_bg = palette.surface
        display_text = palette.text
        button_bg = palette.button
        button_hover = palette.button_hover
        selection_bg = palette.selection
        accent = palette.accent
        accent_hover = palette.accent_hover
        accent_pressed = palette.accent_pressed
        danger = palette.danger
        danger_hover = palette.danger_hover
        
        self.setStyleSheet(f"""
            QWidget {{
                color: {text_color};
                font-family: "{font_family}";
                font-weight: bold;
            }}
            
            QLabel, QPushButton, QComboBox, QCheckBox, QTextEdit, QLineEdit, QSpinBox {{
                font-family: "{font_family}";
                font-weight: bold;
            }}
            
            #leftSidebar {{
                background-color: {sidebar_bg};
                border-bottom-left-radius: 10px;
            }}
            
            #leftSpacer {{
                background-color: {sidebar_bg};
            }}
            
            #rightSidebar {{
                background-color: {sidebar_bg};
                border-bottom-right-radius: 10px;
            }}
            
            #titleBottomSeparator, #leftCenterSeparator, #centerRightSeparator {{
                background-color: {border_color};
                border: none;
            }}
            
            #centerContent {{
                background-color: {content_bg};
            }}
            
            #sidebarTitle {{
                font-size: 15px;
                font-weight: 600;
                color: {text_color};
            }}
            
            #fieldLabel {{
                font-size: 11px;
                font-weight: 500;
                color: {secondary_text};
                text-transform: uppercase;
            }}
            
            QComboBox {{
                background-color: {input_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                padding: 6px 10px;
                padding-right: 30px;
                min-height: 20px;
            }}
            
            QComboBox:hover {{
                border-color: {accent};
            }}

            QLineEdit, QSpinBox {{
                background-color: {input_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                padding: 5px 8px;
                min-height: 20px;
            }}

            QLineEdit:focus, QSpinBox:focus {{
                border-color: {accent};
            }}
            
            QComboBox::drop-down {{
                border: none;
                width: 30px;
                subcontrol-origin: padding;
                subcontrol-position: center right;
            }}
            
            QComboBox::down-arrow {{
                image: url("{resource_path("src/resource/triangle.png").replace(os.sep, "/")}");
                width: 12px;
                height: 12px;
            }}
            
            QComboBox QAbstractItemView {{
                background-color: {palette.elevated};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                selection-background-color: {accent};
                selection-color: white;
            }}
            
            #primaryButton {{
                background-color: {accent};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }}
            
            #primaryButton:hover {{
                background-color: {accent_hover};
            }}
            
            #primaryButton:pressed {{
                background-color: {accent_pressed};
            }}
            
            #primaryButton:disabled {{
                background-color: {border_color};
                color: #86868b;
            }}
            
            #secondaryButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                padding: 6px 12px;
            }}
            
            #secondaryButton:hover {{
                background-color: {button_hover};
            }}
            
            #statusFrame {{
                background-color: {palette.elevated};
                border: none;
                border-radius: 8px;
            }}
            
            #statusDot {{
                color: {palette.inactive};
                font-size: 16px;
            }}
            
            #statusText {{
                color: {secondary_text};
                font-size: 12px;
            }}
            
            #dataDisplayContainer {{
                background-color: {display_bg};
                border: 1px solid {border_color};
                border-radius: 12px;
                padding: 3px;
            }}
            
            #chatScrollArea {{
                background-color: {display_bg};
                border: none;
            }}
            
            #messagesWidget {{
                background-color: {display_bg};
            }}
            
            #chatScrollArea QScrollBar:vertical {{
                background-color: transparent;
                width: 8px;
                border-radius: 4px;
            }}
            
            #chatScrollArea QScrollBar::handle:vertical {{
                background-color: {palette.scroll_handle};
                border-radius: 4px;
                min-height: 20px;
            }}
            
            #chatScrollArea QScrollBar::handle:vertical:hover {{
                background-color: #86868b;
            }}
            
            #dataDisplay {{
                background-color: {display_bg};
                color: {display_text};
                border: none;
                padding: 10px;
                line-height: 1.5;
            }}
            
            #sendInputContainer {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 8px;
                padding: 2px;
            }}
            
            #sendInput {{
                background-color: {input_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 9px;
                padding: 8px 10px;
            }}
            
            #sendInput:focus {{
                border: 1px solid {accent};
                background-color: {palette.elevated};
            }}
            
            #historyList {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                padding: 4px;
            }}
            
            #historyList::item {{
                padding: 6px;
                border-radius: 4px;
            }}
            
            #historyList::item:hover {{
                background-color: {button_hover};
            }}
            
            #historyList::item:selected {{
                background-color: {accent};
                color: white;
            }}
            
            QCheckBox {{
                spacing: 6px;
            }}
            
            QCheckBox::indicator {{
                width: 16px;
                height: 16px;
                border: 1px solid {border_color};
                border-radius: 4px;
                background-color: {input_bg};
            }}
            
            QCheckBox::indicator:checked {{
                background-color: {accent};
                border-color: {accent};
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iMTIiIHZpZXdCb3g9IjAgMCAxMiAxMiIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj4KPHBhdGggZD0iTTEwIDNMNC41IDguNUwyIDYiIHN0cm9rZT0id2hpdGUiIHN0cm9rZS13aWR0aD0iMiIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIiBzdHJva2UtbGluZWpvaW49InJvdW5kIi8+Cjwvc3ZnPgo=);
            }}
            
            #separator {{
                background-color: {border_color};
                max-height: 1px;
            }}
            
            #statsLabel {{
                color: {secondary_text};
                font-size: 12px;
            }}
            
            #darkModeButton, #githubButton {{
                background-color: {button_bg};
                border: 1px solid {border_color};
                border-radius: 16px;
                padding: 6px;
            }}
            
            #darkModeButton:hover, #githubButton:hover {{
                background-color: {border_color};
            }}
            
            #darkModeButton:pressed, #githubButton:pressed {{
                background-color: {border_color};
            }}
            
            /* 快捷输入面板样式 */
            #quickInputHeader {{
                background-color: {sidebar_bg};
                border: 1px solid {border_color};
                border-radius: 6px;
                margin-bottom: 4px;
            }}
            
            #headerLabel {{
                color: {secondary_text};
                font-size: 10px;
                font-weight: 600;
                text-transform: uppercase;
            }}
            
            #quickInputScrollArea {{
                background-color: transparent;
                border: none;
                border-radius: 8px;
            }}
            
            #quickInputCommands {{
                background-color: transparent;
            }}
            
            #quickInputScrollArea QScrollBar:vertical {{
                background-color: transparent;
                width: 8px;
                border-radius: 4px;
            }}
            
            #quickInputScrollArea QScrollBar::handle:vertical {{
                background-color: {palette.scroll_handle};
                border-radius: 4px;
                min-height: 20px;
            }}
            
            #quickInputScrollArea QScrollBar::handle:vertical:hover {{
                background-color: #86868b;
            }}
            
            #commandRow {{
                background-color: {palette.elevated};
                border: 1px solid {palette.separator};
                border-radius: 9px;
                margin: 2px 0px;
            }}
            
            #commandRow:hover {{
                background-color: {selection_bg};
                border-color: {accent};
            }}
            
            #commandContentContainer {{
                background-color: transparent;
                border: none;
            }}
            
            #commandEdit {{
                background-color: transparent;
                color: {text_color};
                border: none;
                padding: 2px 4px;
                font-size: 11px;
            }}
            
            #commandDescriptionEdit {{
                background-color: transparent;
                color: {secondary_text};
                border: none;
                padding: 2px 4px;
                font-size: 9px;
                font-weight: normal;
            }}
            
            #commandDescriptionEdit:focus {{
                color: {text_color};
            }}
            
            #quickSendButton {{
                background-color: {accent};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 0px;
                font-size: 10px;
                font-weight: 500;
            }}
            
            #quickSendButton:hover {{
                background-color: {accent_hover};
            }}
            
            #quickSendButton:pressed {{
                background-color: {accent_pressed};
            }}
            
            #quickSendButton:disabled {{
                background-color: {border_color};
                color: #86868b;
            }}
            
            #deleteButton {{
                background-color: transparent;
                color: {danger};
                border: 1px solid {danger};
                border-radius: 4px;
                padding: 0px;
                font-size: 10px;
                font-weight: 500;
            }}
            
            #deleteButton:hover {{
                background-color: {danger};
                color: white;
            }}
            
            #deleteButton:pressed {{
                background-color: {danger_hover};
            }}
            
            #addCommandButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 11px;
                font-weight: 500;
            }}
            
            #addCommandButton:hover {{
                background-color: {button_hover};
            }}
            
            #addCommandButton:pressed {{
                background-color: {palette.separator};
            }}
            
            /* 分区样式 */
            #sectionTitle {{
                color: {text_color};
                font-size: 13px;
                font-weight: 600;
                margin: 4px 0px;
            }}

            #hintLabel {{
                color: {secondary_text};
                background: transparent;
            }}

            #conversionError {{
                color: {danger};
                background: transparent;
            }}

            QLineEdit#radixDisplay {{
                background: transparent;
                color: {text_color};
                border: none;
                border-bottom: 1px solid {palette.separator};
                border-radius: 0px;
                padding: 10px 4px 12px 4px;
                font-size: 28px;
                font-weight: 500;
                selection-background-color: {selection_bg};
            }}

            QFrame#baseValueRow {{
                background: transparent;
                border: none;
                border-radius: 8px;
            }}

            QFrame#baseValueRow:hover {{
                background-color: {button_hover};
            }}

            QFrame#baseSelectionIndicator {{
                background-color: {accent};
                border: none;
                border-radius: 2px;
            }}

            QLabel#baseName {{
                color: {secondary_text};
                background: transparent;
                font-size: 13px;
            }}

            QLabel#baseValue {{
                color: {text_color};
                background: transparent;
                font-size: 13px;
            }}

            QTableWidget#asciiTable {{
                background-color: {input_bg};
                alternate-background-color: {bg_color};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 8px;
                gridline-color: {palette.separator};
                selection-background-color: {selection_bg};
                selection-color: {text_color};
            }}

            QTableWidget#asciiTable QHeaderView::section {{
                background-color: {button_bg};
                color: {secondary_text};
                border: none;
                border-bottom: 1px solid {border_color};
                padding: 6px 4px;
            }}
            
            #sectionSeparator {{
                background-color: {border_color};
                border: none;
                margin: 8px 0px;
            }}
            
            QSplitter#workbenchSplitter::handle {{
                background-color: transparent;
                border: none;
                image: none;
                margin: 12px 3px;
                border-radius: 1px;
            }}

            QSplitter#workbenchSplitter::handle:hover {{
                background-color: {palette.separator};
            }}

            QPushButton#edgeToggleButton {{
                background-color: {palette.elevated};
                color: {secondary_text};
                border: 1px solid {palette.separator};
                border-radius: 10px;
                padding: 0px;
                font-size: 19px;
                font-weight: 500;
            }}

            QPushButton#edgeToggleButton:hover {{
                background-color: {selection_bg};
                color: {accent};
                border-color: {accent};
            }}

            QPushButton#edgeToggleButton:pressed {{
                background-color: {button_hover};
            }}

            #sendComposer {{
                background-color: {palette.elevated};
                border: 1px solid {palette.separator};
                border-radius: 12px;
            }}

            #composerUtilityButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 9px;
                padding: 0px 8px;
            }}

            #composerUtilityButton:hover {{
                background-color: {button_hover};
                border-color: {accent};
            }}

            #payloadStatus {{
                color: {secondary_text};
                background: transparent;
            }}

            #payloadStatusError {{
                color: {palette.danger};
                background: transparent;
            }}

            #toolbarButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 7px;
                padding: 5px 9px;
            }}

            #toolbarButton:hover {{
                background-color: {button_hover};
            }}

            #toolbarButton:checked {{
                background-color: {accent};
                color: white;
                border-color: {accent};
            }}

            QTabWidget#inspectorTabs::pane {{
                border: none;
                background: transparent;
            }}

            QWidget#inspectorPage {{
                background-color: {sidebar_bg};
            }}

            QTabBar#inspectorTabBar {{
                background: transparent;
                border: none;
            }}

            QTabBar#inspectorTabBar::tab {{
                background: transparent;
                color: {secondary_text};
                border: none;
                padding: 7px 14px;
                margin-right: 2px;
            }}

            QTabBar#inspectorTabBar::tab:selected {{
                color: {accent};
                background-color: {selection_bg};
                border-radius: 7px;
            }}

            #lineControlCard {{
                background-color: {palette.elevated};
                border: 1px solid {palette.separator};
                border-radius: 10px;
            }}

            #connectionSummary {{
                color: {secondary_text};
                font-size: 10px;
                font-weight: 500;
            }}

            #lineStateInactive, #lineStateActive {{
                border-radius: 6px;
                padding: 4px 2px;
                font-size: 9px;
            }}

            #lineStateInactive {{
                color: {tertiary_text};
                background-color: {button_bg};
            }}

            #lineStateActive {{
                color: {palette.success};
                background-color: {selection_bg};
            }}

            QPushButton#lineControlButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 7px;
                padding: 5px 4px;
            }}

            QPushButton#lineControlButton:hover {{
                background-color: {button_hover};
            }}

            QPushButton#lineControlButton:checked {{
                background-color: {accent};
                border-color: {accent};
                color: white;
            }}

            QPushButton#lineControlButton:disabled {{
                color: {tertiary_text};
                background-color: {button_bg};
                border-color: {palette.separator};
            }}

            /* 紧凑标签样式 */
            #compactLabel {{
                color: {text_color};
                font-size: 11px;
                font-weight: 500;
                margin: 2px 0px;
            }}
            
            /* 信息标签样式 */
            #infoLabel {{
                color: {secondary_text};
                font-size: 9px;
                font-weight: normal;
                margin: 2px 0px;
            }}
            
            /* 刷新按钮样式 */
            #refreshButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                font-size: 16px;
                font-weight: bold;
            }}
            
            #refreshButton:hover {{
                background-color: {button_hover};
            }}
            
            #refreshButton:pressed {{
                background-color: {palette.separator};
            }}
            
            /* SpinBox 样式 */
            QSpinBox {{
                background-color: {input_bg};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                padding: 6px 10px;
                min-height: 20px;
            }}
            
            QSpinBox:hover {{
                border-color: {accent};
            }}
            
            QSpinBox::up-button, QSpinBox::down-button,
            QAbstractSpinBox::up-button, QAbstractSpinBox::down-button {{
                border: none;
                width: 0px;
                height: 0px;
                subcontrol-position: top right;
            }}

            QSpinBox::up-arrow, QSpinBox::down-arrow,
            QAbstractSpinBox::up-arrow, QAbstractSpinBox::down-arrow {{
                image: none;
                width: 0px;
                height: 0px;
            }}
        """)

    def update_connection_summary(self):
        """Refresh the compact serial format summary shown in the sidebar."""
        if self.current_mode != 'serial':
            return
        parity = {
            "None": "N", "Even": "E", "Odd": "O", "Mark": "M", "Space": "S"
        }.get(self.parity_combo.currentText(), "N")
        flow = self.flow_control_combo.currentText()
        flow_text = "无流控" if flow == "None" else flow
        self.connection_summary.setText(
            f"{self.baud_combo.currentText()} · "
            f"{self.data_bits_combo.currentText()}{parity}{self.stop_bits_combo.currentText()}"
            f" · {flow_text}"
        )

    def update_line_states(self):
        """Poll modem lines at a low rate and update only their visual state."""
        connected = self.current_mode == 'serial' and self.serial_manager.is_connected
        states = self.serial_manager.line_states if connected else {}
        for name, label in self.line_state_labels.items():
            active = bool(states.get(name, False))
            label.setProperty("active", active)
            label.setObjectName("lineStateActive" if active else "lineStateInactive")
            label.style().unpolish(label)
            label.style().polish(label)
        for button in (self.dtr_button, self.rts_button, self.break_button):
            button.setEnabled(connected)

    def set_dtr_line(self, enabled: bool):
        if self.serial_manager.is_connected and not self.serial_manager.set_dtr(enabled):
            self.dtr_button.setChecked(not enabled)

    def set_rts_line(self, enabled: bool):
        if self.serial_manager.is_connected and not self.serial_manager.set_rts(enabled):
            self.rts_button.setChecked(not enabled)

    def send_break_signal(self):
        if self.serial_manager.is_connected:
            self.serial_manager.send_break(0.25)
    
    def on_mode_changed(self, mode_text: str):
        """模式切换"""
        # 如果已连接，先断开
        if self.is_any_connected():
            self.toggle_connection()
        
        if mode_text == "串口":
            self.current_mode = 'serial'
            self.serial_config_widget.show()
            self.bluetooth_config_widget.hide()
            self.refresh_btn.setToolTip("刷新串口列表")
            self.line_control_card.show()
        else:  # 蓝牙
            self.current_mode = 'bluetooth'
            self.serial_config_widget.hide()
            self.bluetooth_config_widget.show()
            self.refresh_btn.setToolTip("扫描蓝牙设备")
            self.line_control_card.hide()
            
            # 检查蓝牙是否可用
            if not BluetoothManager.is_available():
                self.show_macos_alert(
                    "蓝牙不可用",
                    "蓝牙功能需要安装蓝牙库\n\n请运行: pip install bleak"
                )
            else:
                # 根据蓝牙后端显示不同的提示
                backend = BluetoothManager.get_backend()
                if backend == 'bleak':
                    self.bt_info_label.setText("BLE 模式\n自动检测设备特征")
                    self.pybluez_config_widget.hide()
                elif backend == 'pybluez':
                    self.bt_info_label.setText("经典蓝牙模式\n需要先配对设备")
                    self.pybluez_config_widget.show()
        
        self.refresh_devices()
        self.update_connection_summary()
    
    def is_any_connected(self) -> bool:
        """检查是否有任何连接"""
        if self.current_mode == 'serial':
            return self.serial_manager.is_connected
        else:
            return self.bluetooth_manager.is_connected
    
    def refresh_devices(self):
        """刷新设备列表"""
        if self.current_mode == 'serial':
            self.refresh_ports()
        else:
            self.scan_bluetooth_devices()
    
    def scan_bluetooth_devices(self):
        """扫描蓝牙设备"""
        if not BluetoothManager.is_available():
            return
        
        # 禁用刷新按钮，显示扫描中
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.setText("...")
        self.port_combo.clear()
        self.port_combo.addItem("正在扫描...")
        
        # 在后台线程中扫描
        def scan_thread():
            devices = self.bluetooth_manager.scan_devices(duration=8)
            # 使用信号更新UI
            self.bluetooth_manager.scan_completed.emit(devices)
        
        import threading
        thread = threading.Thread(target=scan_thread, daemon=True)
        thread.start()
    
    def on_bluetooth_scan_completed(self, devices: List[BluetoothDevice]):
        """蓝牙扫描完成"""
        self.refresh_btn.setEnabled(True)
        self.refresh_btn.setText("⟳")
        self.port_combo.clear()
        
        if not devices:
            self.port_combo.addItem("未发现设备")
        else:
            for device in devices:
                self.port_combo.addItem(device.display_name, device.address)
    
    def connect_signals(self):
        """连接信号"""
        # 串口信号
        self.serial_manager.data_received.connect(self.on_data_received)
        self.serial_manager.device_connected.connect(self.on_device_connected)
        self.serial_manager.device_disconnected.connect(self.on_device_disconnected)
        self.serial_manager.error_occurred.connect(self.on_error)
        self.serial_manager.connecting_status.connect(self.on_connecting_status)
        
        # 蓝牙信号
        self.bluetooth_manager.data_received.connect(self.on_data_received)
        self.bluetooth_manager.device_connected.connect(self.on_device_connected)
        self.bluetooth_manager.device_disconnected.connect(self.on_device_disconnected)
        self.bluetooth_manager.error_occurred.connect(self.on_error)
        self.bluetooth_manager.scan_completed.connect(self.on_bluetooth_scan_completed)
        self.bluetooth_manager.connecting_status.connect(self.on_connecting_status)
        
        # 内部信号
        self.display_sent_signal.connect(self.display_sent_data)
        
        # 串口配置变更信号 - 支持动态修改
        self.baud_combo.currentTextChanged.connect(self.on_serial_config_changed)
        self.data_bits_combo.currentTextChanged.connect(self.on_serial_config_changed)
        self.stop_bits_combo.currentTextChanged.connect(self.on_serial_config_changed)
        self.parity_combo.currentTextChanged.connect(self.on_serial_config_changed)
        self.flow_control_combo.currentTextChanged.connect(self.on_serial_config_changed)
        self.port_combo.currentTextChanged.connect(self.update_connection_summary)
    
    def refresh_ports(self):
        """刷新串口列表"""
        self.port_combo.clear()
        devices = SerialManager.list_devices()
        for device in devices:
            self.port_combo.addItem(device.display_name, device.port)
    
    def on_serial_config_changed(self):
        """串口配置变更 - 实时应用到已连接的串口"""
        try:
            config = {
                'baud_rate': int(self.baud_combo.currentText()),
                'data_bits': int(self.data_bits_combo.currentText()),
                'stop_bits': float(self.stop_bits_combo.currentText()),
                'parity': self.parity_combo.currentText(),
                'flow_control': self.flow_control_combo.currentText(),
            }
            for key, value in config.items():
                self.config.set(f"serial.{key}", value)
            self.config.save_config()
            if self.current_mode == 'serial' and self.serial_manager.is_connected:
                self.serial_manager.configure(config)
            self.update_connection_summary()
        except ValueError:
            # Editable baud input can be temporarily incomplete while typing.
            pass

    def set_auto_reconnect(self, enabled):
        enabled = bool(enabled)
        self.serial_manager.set_auto_reconnect(enabled)
        self.config.set("serial.auto_reconnect", enabled)
        self.config.save_config()
    
    def toggle_connection(self):
        """切换连接状态"""
        if self.current_mode == 'serial':
            # 串口模式
            if self.serial_manager.is_connected:
                self.serial_manager.disconnect()
            else:
                port = self.port_combo.currentData()
                if not port:
                    self.on_error("请选择串口")
                    return
                
                # 配置串口参数
                config = {
                    'baud_rate': int(self.baud_combo.currentText()),
                    'data_bits': int(self.data_bits_combo.currentText()),
                    'stop_bits': float(self.stop_bits_combo.currentText()),
                    'parity': self.parity_combo.currentText(),
                    'flow_control': self.flow_control_combo.currentText(),
                }
                self.serial_manager.configure(config)
                
                # 在后台线程中连接，避免阻塞 UI
                import threading
                def connect_thread():
                    self.serial_manager.connect(port)
                
                thread = threading.Thread(target=connect_thread, daemon=True)
                thread.start()
        else:
            # 蓝牙模式
            if self.bluetooth_manager.is_connected:
                self.bluetooth_manager.disconnect()
            else:
                address = self.port_combo.currentData()
                if not address:
                    self.on_error("请选择蓝牙设备")
                    return
                
                port = self.rfcomm_port_spin.value()
                
                # 在后台线程中连接，避免阻塞 UI
                import threading
                def connect_thread():
                    self.bluetooth_manager.connect(address, port)
                
                thread = threading.Thread(target=connect_thread, daemon=True)
                thread.start()
    
    def send_data(self):
        """Validate once, then send the exact analyzed bytes."""
        manager = self.serial_manager if self.current_mode == 'serial' else self.bluetooth_manager

        if not manager.is_connected:
            self.on_error("设备未连接")
            return

        text = self.send_input.toPlainText()
        if not text:
            return
        format_type = self.send_format_combo.currentText()
        analysis = analyze_payload(
            text,
            format_type,
            self.add_carriage.isChecked(),
            self.add_newline.isChecked(),
        )
        if not analysis.valid:
            self.on_error(analysis.error)
            self.update_send_analysis()
            return
        if not analysis.payload:
            return

        display_text = analysis.normalized if format_type == "HEX" else text
        self.display_sent_data(display_text, format_type, analysis.payload)
        self.send_history.add(text, format_type)

        if not self.repeat_check.isChecked():
            self.send_input.clear()

        import threading

        def send_thread():
            success = manager.send_data(analysis.payload)
            self.manual_send_completed.emit(success, analysis.byte_count)

        thread = threading.Thread(target=send_thread, daemon=True)
        thread.start()

    def _on_manual_send_completed(self, success: bool, byte_count: int):
        if success:
            self.sent_count += 1
            self.sent_bytes += byte_count
            self.update_stats()

    def run_enabled_commands(self):
        manager = self.serial_manager if self.current_mode == "serial" else self.bluetooth_manager
        if not manager.is_connected:
            self.on_error("设备未连接")
            return
        if self.sequence_runner.start(
            self.quick_commands,
            self.send_format_combo.currentText(),
            self._send_sequence_command,
        ):
            self.run_sequence_btn.setEnabled(False)
            self.stop_sequence_btn.setEnabled(True)

    def stop_command_sequence(self):
        self.sequence_runner.stop()

    def _send_sequence_command(self, command: str, format_type: str) -> bool:
        manager = self.serial_manager if self.current_mode == "serial" else self.bluetooth_manager
        if format_type == "HEX":
            return manager.send_hex_string(command)
        encoding = "ascii" if format_type == "ASCII" else "utf-8"
        return manager.send_text(command, encoding)

    def _on_sequence_command_sent(self, command: str, format_type: str):
        self.display_sent_data(command, format_type)

    def _on_sequence_command_result(self, index: int, success: bool, byte_count: int):
        if success:
            self.sent_count += 1
            self.sent_bytes += byte_count
            self.update_stats()

    def _on_sequence_completed(self):
        self.run_sequence_btn.setEnabled(True)
        self.stop_sequence_btn.setEnabled(False)

    def toggle_repeat_send(self, state):
        """Start or stop repeated sending without blocking the UI thread."""
        if state == Qt.CheckState.Checked.value:
            if not self.send_input.toPlainText().strip():
                self.repeat_check.setChecked(False)
                self.on_error("请先输入要循环发送的数据")
                return
            self.repeat_timer.start(1000)
        else:
            self.repeat_timer.stop()

    def update_repeat_interval(self, interval: int):
        if self.repeat_timer.isActive():
            self.repeat_timer.start(interval)
    
    def display_sent_data(self, text: str, format_type: str, payload: bytes = None):
        """在数据显示区显示发送的数据 - QQ聊天样式（右对齐，蓝色气泡）"""
        if payload is None:
            self._record_payload("发送", text, format_type)
        else:
            self.session_recorder.record("发送", payload)
        # 格式化显示
        if format_type == "HEX":
            try:
                hex_string = ''.join(c for c in text if c in '0123456789ABCDEFabcdef')
                if len(hex_string) % 2 != 0:
                    hex_string = '0' + hex_string
                data = bytes.fromhex(hex_string)
                display = ' '.join(f'{b:02X}' for b in data)
            except:
                display = text
        else:
            display = text
        
        try:
            # 获取时间戳
            from datetime import datetime
            timestamp = datetime.now().strftime("%H:%M:%S")
            
            # 创建气泡消息
            self.add_message_bubble(display, timestamp, is_sent=True)
        except Exception as e:
            import traceback
            traceback.print_exc()
    
    def _record_payload(self, direction: str, text: str, format_type: str):
        """Convert displayed text back to bytes for lossless session logging."""
        try:
            if format_type == "HEX":
                hex_text = "".join(c for c in text if c in "0123456789abcdefABCDEF")
                if len(hex_text) % 2:
                    hex_text = "0" + hex_text
                payload = bytes.fromhex(hex_text)
            elif format_type == "ASCII":
                payload = text.encode("ascii", errors="replace")
            else:
                payload = text.encode("utf-8")
            self.session_recorder.record(direction, payload)
        except (TypeError, ValueError, UnicodeError):
            pass

    def toggle_recording(self):
        """Start or stop a CSV session recording."""
        if self.session_recorder.is_recording:
            path = self.session_recorder.stop()
            self.record_btn.setText("录制")
            self.record_btn.setToolTip("将收发原始数据保存为 CSV")
            if path:
                self.show_macos_alert("录制完成", f"会话已保存到：\n{path}")
            return

        default_name = datetime.now().strftime("serial_session_%Y%m%d_%H%M%S.csv")
        path, _ = QFileDialog.getSaveFileName(self, "保存会话记录", default_name, "CSV 文件 (*.csv)")
        if path and self.session_recorder.start(path):
            self.record_btn.setText("停止录制")
            self.record_btn.setToolTip("停止并保存当前会话")

    def filter_messages(self, query: str):
        """Filter message bubbles without deleting recorded data."""
        normalized_query = query.casefold().strip()
        for index in range(self.messages_layout.count()):
            item = self.messages_layout.itemAt(index)
            widget = item.widget() if item else None
            if widget is None or widget.property("message_text") is None:
                continue
            message_text = str(widget.property("message_text"))
            widget.setVisible(not normalized_query or normalized_query in message_text)

    def toggle_timestamps(self):
        """切换时间戳显示"""
        show_timestamps = self.timestamp_check.isChecked()
        
        # 遍历所有消息容器，更新时间戳显示
        for i in range(self.messages_layout.count()):
            item = self.messages_layout.itemAt(i)
            if item and item.widget():
                message_container = item.widget()
                # 查找时间戳标签
                for child in message_container.findChildren(QLabel):
                    if child.objectName() == "timestampLabel":
                        child.setVisible(show_timestamps)
    
    def add_message_bubble(self, text: str, timestamp: str, is_sent: bool = True):
        """添加消息气泡"""
        # 记录消息用于导出
        self.message_log.append({
            "time": timestamp,
            "direction": "发送" if is_sent else "接收",
            "content": text,
        })

        # 创建消息容器（包含时间戳和气泡）
        message_container = MessageContainer()
        message_container.setProperty("message_text", text.casefold())
        container_layout = QVBoxLayout(message_container)
        container_layout.setContentsMargins(0, 4, 0, 4)
        container_layout.setSpacing(2)
        
        # 时间戳（在气泡上方）
        time_label = QLabel(timestamp)
        time_label.setObjectName("timestampLabel")
        time_label.setAlignment(Qt.AlignmentFlag.AlignRight if is_sent else Qt.AlignmentFlag.AlignLeft)
        
        if self.timestamp_check.isChecked():
            # 显示时间戳
            time_label.setStyleSheet("""
                color: #86868b;
                font-size: 11px;
                background: transparent;
                padding: 2px 8px;
            """)
            time_label.setVisible(True)
        else:
            # 隐藏时间戳
            time_label.setVisible(False)
        
        container_layout.addWidget(time_label)
        
        # 气泡容器（用于左右对齐）
        bubble_container = QWidget()
        bubble_layout = QHBoxLayout(bubble_container)
        bubble_layout.setContentsMargins(0, 0, 0, 0)
        bubble_layout.setSpacing(0)
        
        # 创建气泡
        bubble = QFrame()
        bubble.setObjectName("sentBubble" if is_sent else "receivedBubble")
        bubble_content_layout = QVBoxLayout(bubble)
        bubble_content_layout.setContentsMargins(14, 10, 14, 10)
        bubble_content_layout.setSpacing(0)
        
        # 消息内容
        content_label = QLabel(text)
        content_label.setObjectName("bubbleContent")
        content_label.setWordWrap(True)
        content_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        bubble_content_layout.addWidget(content_label)
        
        # 设置气泡样式和对齐
        if is_sent:
            # 发送消息：右对齐，蓝色
            bubble_layout.addStretch()
            bubble_layout.addWidget(bubble)
            bubble.setStyleSheet("""
                #sentBubble {
                    background-color: #0071e3;
                    border-radius: 18px;
                    max-width: 400px;
                    min-width: 60px;
                }
                #bubbleContent {
                    color: #ffffff;
                    font-size: 14px;
                    background: transparent;
                }
            """)
        else:
            # 接收消息：左对齐，灰色
            bubble_layout.addWidget(bubble)
            bubble_layout.addStretch()
            bubble_color = "#e5e5ea" if not self.is_dark_mode else "#3a3a3c"
            text_color = "#000000" if not self.is_dark_mode else "#ffffff"
            bubble.setStyleSheet(f"""
                #receivedBubble {{
                    background-color: {bubble_color};
                    border-radius: 18px;
                    max-width: 400px;
                    min-width: 60px;
                }}
                #bubbleContent {{
                    color: {text_color};
                    font-size: 14px;
                    background: transparent;
                }}
            """)
        
        container_layout.addWidget(bubble_container)

        # 复制图标行：固定高度避免悬停时布局抖动
        copy_row = QWidget()
        copy_row.setFixedHeight(20)
        copy_row_layout = QHBoxLayout(copy_row)
        copy_row_layout.setContentsMargins(4, 0, 4, 0)
        copy_row_layout.setSpacing(0)

        copy_btn = QPushButton(copy_row)
        copy_btn.setFixedSize(16, 16)
        copy_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        icon_path = resource_path("src/resource/copy.png")
        copy_btn.setIcon(QIcon(icon_path))
        copy_btn.setIconSize(QSize(13, 13))
        copy_btn.setStyleSheet("QPushButton { background: transparent; border: none; }")
        copy_btn.setVisible(False)

        copied_label = QLabel("已复制", copy_row)
        copied_label.setStyleSheet("color: #34c759; font-size: 11px; background: transparent;")
        copied_label.setVisible(False)

        # 把 text 存到按钮属性上，避免闭包问题
        copy_btn.setProperty("msg_text", text)
        copy_btn.clicked.connect(self._on_copy_clicked)

        if is_sent:
            copy_row_layout.addStretch()
            copy_row_layout.addWidget(copied_label)
            copy_row_layout.addWidget(copy_btn)
        else:
            copy_row_layout.addWidget(copy_btn)
            copy_row_layout.addWidget(copied_label)
            copy_row_layout.addStretch()

        # 把 copy_btn / copied_label 存到 message_container 上供 hover 使用
        message_container.set_copy_widgets(copy_btn, copied_label)

        container_layout.addWidget(copy_row)

        # 插入到消息列表（在stretch之前）
        count = self.messages_layout.count()
        self.messages_layout.insertWidget(count - 1, message_container)

        max_records = max(100, int(self.config.get("display.max_records", 1000)))
        while len(self.message_log) > max_records:
            self.message_log.pop(0)
            oldest = self.messages_layout.takeAt(0)
            if oldest and oldest.widget():
                oldest.widget().deleteLater()
        
        # 自动滚动到底部
        if self.autoscroll_check.isChecked():
            QTimer.singleShot(50, lambda: self.scroll_area.verticalScrollBar().setValue(
                self.scroll_area.verticalScrollBar().maximum()
            ))

    def _on_copy_clicked(self):
        """复制按钮点击处理"""
        btn = self.sender()
        if btn is None:
            return
        msg_text = btn.property("msg_text")
        if msg_text:
            QApplication.clipboard().setText(msg_text)
        # 找到同级的 copied_label（btn 和 label 在同一个 copy_row 里）
        copy_row = btn.parent()
        if copy_row is None:
            return
        copied_label = None
        for child in copy_row.children():
            if isinstance(child, QLabel):
                copied_label = child
                break
        if copied_label is None:
            return
        btn.setVisible(False)
        copied_label.setVisible(True)
        QTimer.singleShot(1500, lambda: (
            copied_label.setVisible(False),
            btn.setVisible(True),
        ))
    
    
    def export_to_excel(self):
        """导出消息记录到 Excel"""
        if not self.message_log:
            self.show_macos_alert("提示", "没有可导出的消息记录")
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            self.show_macos_alert("缺少依赖", "请先安装 openpyxl：\npip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", "消息记录.xlsx", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "消息记录"

        # 表头
        headers = ["时间", "方向", "内容"]
        ws.append(headers)
        header_fill_sent = PatternFill("solid", fgColor="0071E3")
        header_font = Font(bold=True, color="FFFFFF")
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill_sent
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 数据行
        sent_fill = PatternFill("solid", fgColor="D6EAFF")
        recv_fill = PatternFill("solid", fgColor="F2F2F7")
        for msg in self.message_log:
            ws.append([msg["time"], msg["direction"], msg["content"]])
            row = ws.max_row
            fill = sent_fill if msg["direction"] == "发送" else recv_fill
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="center")

        # 列宽
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 60

        wb.save(path)
        self.show_macos_alert("导出成功", f"已保存到：\n{path}")

    def toggle_receive_pause(self, paused: bool):
        """Pause rendering while keeping acquisition and recording active."""
        self.receive_pause_buffer.set_paused(paused)
        if paused:
            self.pause_display_btn.setText("继续")
            self.pause_display_btn.setToolTip("继续显示已缓存的接收数据")
            return

        data, dropped = self.receive_pause_buffer.drain()
        self.pause_display_btn.setText("暂停显示")
        self.pause_display_btn.setToolTip("暂停界面刷新，但继续接收并记录数据")
        if dropped:
            self.add_message_bubble(
                f"暂停期间缓冲区已满，已丢弃 {dropped} 字节的早期显示数据",
                datetime.now().strftime("%H:%M:%S"),
                is_sent=False,
            )
        if data:
            self._display_received_data(data)

    def _update_pause_button(self):
        byte_count = self.receive_pause_buffer.byte_count
        if byte_count < 1024:
            size = f"{byte_count} B"
        elif byte_count < 1024 * 1024:
            size = f"{byte_count / 1024:.1f} KB"
        else:
            size = f"{byte_count / (1024 * 1024):.1f} MB"
        self.pause_display_btn.setText(f"继续 · {size}")

    def clear_display(self):
        """清除显示"""
        # 清除所有消息气泡
        while self.messages_layout.count() > 1:  # 保留最后的stretch
            item = self.messages_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self.sent_count = 0
        self.sent_bytes = 0
        self.received_count = 0
        self.received_bytes = 0
        self.message_log.clear()
        self.receive_pause_buffer.clear()
        if self.pause_display_btn.isChecked():
            self._update_pause_button()
        self.update_stats()
    
    def on_data_received(self, data: bytes):
        """接收到数据 - QQ聊天样式（左对齐，灰色气泡）"""
        self.session_recorder.record("接收", data)
        self.received_count += 1
        self.received_bytes += len(data)
        self.update_stats()

        if not self.receive_pause_buffer.append(data):
            self._update_pause_button()
            return
        self._display_received_data(data)

    def _display_received_data(self, data: bytes):
        """Format and render received bytes without changing acquisition stats."""
        format_type = self.format_combo.currentText()
        if format_type == "HEX":
            text = ' '.join(f'{b:02X}' for b in data)
        elif format_type == "ASCII":
            text = data.decode('ascii', errors='replace')
        else:  # UTF-8
            text = data.decode('utf-8', errors='replace')

        try:
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.add_message_bubble(text, timestamp, is_sent=False)
        except Exception:
            import traceback
            traceback.print_exc()
    
    def on_connecting_status(self, status: str):
        """连接状态更新"""
        if status:
            # 开始连接动画
            self.connecting_dots = 0
            self.status_label.setText(status)
            self.status_indicator.setStyleSheet(
                f"color: {palette_for(self.is_dark_mode).warning};"
            )
            self.connecting_timer.start(500)  # 每500ms更新一次
            
            # 禁用连接按钮
            self.connect_btn.setEnabled(False)
        else:
            # 停止连接动画
            self.connecting_timer.stop()
            self.connect_btn.setEnabled(True)
    
    def update_connecting_animation(self):
        """更新连接动画"""
        self.connecting_dots = (self.connecting_dots + 1) % 4
        dots = "." * self.connecting_dots
        self.status_label.setText(f"正在连接{dots}")
    
    def on_device_connected(self):
        """设备已连接"""
        self.connecting_timer.stop()  # 停止连接动画
        self.connect_btn.setText("断开")
        self.connect_btn.setEnabled(True)
        self.status_indicator.setStyleSheet(
            f"color: {palette_for(self.is_dark_mode).success};"
        )
        
        if self.current_mode == 'serial':
            self.status_label.setText("串口已连接")
            # 保持串口配置控件启用，允许动态修改
            self.port_combo.setEnabled(False)  # 只禁用端口选择
            self.baud_combo.setEnabled(True)
            self.data_bits_combo.setEnabled(True)
            self.stop_bits_combo.setEnabled(True)
            self.parity_combo.setEnabled(True)
        else:
            self.status_label.setText("蓝牙已连接")
            # 禁用蓝牙配置控件
            self.port_combo.setEnabled(False)
            self.rfcomm_port_spin.setEnabled(False)
        
        # 禁用模式切换
        self.mode_combo.setEnabled(False)
        self.refresh_btn.setEnabled(False)
        
        # 启用发送控件
        self.send_btn.setEnabled(True)
        self.update_line_states()
    
    def on_device_disconnected(self):
        """设备已断开"""
        self.connecting_timer.stop()  # 停止连接动画
        self.connect_btn.setText("连接")
        self.connect_btn.setEnabled(True)
        self.status_indicator.setStyleSheet(
            f"color: {palette_for(self.is_dark_mode).inactive};"
        )
        self.status_label.setText("未连接")
        
        if self.current_mode == 'serial':
            # 启用串口配置控件
            self.port_combo.setEnabled(True)
            self.baud_combo.setEnabled(True)
            self.data_bits_combo.setEnabled(True)
            self.stop_bits_combo.setEnabled(True)
            self.parity_combo.setEnabled(True)
        else:
            # 启用蓝牙配置控件
            self.port_combo.setEnabled(True)
            self.rfcomm_port_spin.setEnabled(True)
        
        # 启用模式切换
        self.mode_combo.setEnabled(True)
        self.refresh_btn.setEnabled(True)
        
        # 禁用发送控件
        self.send_btn.setEnabled(False)
        with QSignalBlocker(self.dtr_button):
            self.dtr_button.setChecked(False)
        with QSignalBlocker(self.rts_button):
            self.rts_button.setChecked(False)
        self.update_line_states()
    
    def on_error(self, error_msg: str):
        """错误处理"""
        self.show_macos_alert("提示", error_msg)
    
    def update_stats(self):
        """更新统计信息"""
        self.sent_stats.setText(f"发送: {self.sent_count} ({self.sent_bytes} 字节)")
        self.received_stats.setText(f"接收: {self.received_count} ({self.received_bytes} 字节)")
        now = time.monotonic()
        elapsed = now - self._rate_timestamp
        if elapsed >= 0.25:
            self._send_rate = (self.sent_bytes - self._rate_sent_bytes) / elapsed
            self._receive_rate = (self.received_bytes - self._rate_received_bytes) / elapsed
            self._rate_timestamp = now
            self._rate_sent_bytes = self.sent_bytes
            self._rate_received_bytes = self.received_bytes
            self.rate_stats.setText(
                f"速率: ↓ {self._receive_rate:.0f} B/s  ↑ {self._send_rate:.0f} B/s"
            )
    
    def toggle_dark_mode(self):
        """切换黑夜模式"""
        self.is_dark_mode = not self.is_dark_mode
        self.apply_macos_style()  # 重新应用样式
        if self.serial_manager.is_connected:
            self.status_indicator.setStyleSheet(
                f"color: {palette_for(self.is_dark_mode).success};"
            )
        else:
            self.status_indicator.setStyleSheet(
                f"color: {palette_for(self.is_dark_mode).inactive};"
            )
        
        # 通知主窗口更新样式
        main_window = self.parent()
        while main_window and not hasattr(main_window, 'apply_window_style'):
            main_window = main_window.parent()
        if main_window:
            main_window.update_theme(self.is_dark_mode)
    
    def open_github(self):
        """打开 GitHub 仓库"""
        import webbrowser
        webbrowser.open("https://github.com/suci135/Suci-Serial-Port-Assistant")
    
    def show_macos_alert(self, title: str, message: str):
        """显示 macOS 风格的提示框"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QGraphicsDropShadowEffect
        from PyQt6.QtCore import Qt, QRect
        from PyQt6.QtGui import QRegion, QPainterPath
        
        dialog = QDialog(self)
        dialog.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        dialog.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        dialog.setModal(True)
        
        # 主容器
        main_container = QWidget()
        main_container.setObjectName("alertContainer")
        main_container.setAutoFillBackground(True)
        main_container.setStyleSheet("""
            #alertContainer {{
                background-color: #ffffff;
                border-radius: 16px;
                border: 1px solid #d2d2d7;
            }}
            QWidget {{
                background-color: #ffffff;
            }}
            #alertTitle {{
                font-size: 14px;
                font-weight: 600;
                color: #1d1d1f;
                background-color: transparent;
            }}
            #alertMessage {{
                font-size: 12px;
                color: #1d1d1f;
                background-color: transparent;
            }}
            #alertButton {{
                background-color: #0071e3;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 28px;
                font-weight: 500;
                min-width: 80px;
            }}
            #alertButton:hover {{
                background-color: #0077ed;
            }}
            #alertButton:pressed {{
                background-color: #006edb;
            }}
        """)
        
        container_layout = QVBoxLayout(main_container)
        container_layout.setContentsMargins(24, 20, 24, 20)
        container_layout.setSpacing(16)
        
        # 标题
        title_label = QLabel(title)
        title_label.setObjectName("alertTitle")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(title_label)
        
        # 消息
        message_label = QLabel(message)
        message_label.setObjectName("alertMessage")
        message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        message_label.setWordWrap(True)
        message_label.setMinimumWidth(250)
        container_layout.addWidget(message_label)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        ok_button = QPushButton("确定")
        ok_button.setObjectName("alertButton")
        ok_button.clicked.connect(dialog.accept)
        ok_button.setCursor(Qt.CursorShape.PointingHandCursor)
        button_layout.addWidget(ok_button)
        button_layout.addStretch()
        container_layout.addLayout(button_layout)
        
        # 设置对话框布局
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.setContentsMargins(10, 10, 10, 10)
        dialog_layout.addWidget(main_container)
        
        # 添加阴影效果
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 60))
        shadow.setOffset(0, 4)
        main_container.setGraphicsEffect(shadow)
        
        dialog.exec()
    
    def load_quick_commands(self):
        """从配置文件加载快捷命令"""
        return self.quick_command_store.load()
    
    def save_quick_commands(self):
        """保存快捷命令到配置文件"""
        self.quick_command_store.save(self.quick_commands)

    def closeEvent(self, event):
        self.repeat_timer.stop()
        self.session_recorder.stop()
        super().closeEvent(event)

    def resizeEvent(self, event):
        """Keep the center toolbar readable at compact window sizes."""
        width = self.width()
        compact = width < 1050
        very_compact = width < 960
        self.search_input.setVisible(not compact)
        self.export_btn.setVisible(not compact)
        self.autoscroll_check.setVisible(not very_compact)
        self.clear_btn.setText("清" if very_compact else "清除")
        super().resizeEvent(event)
